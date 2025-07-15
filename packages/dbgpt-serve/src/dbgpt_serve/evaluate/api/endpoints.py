import logging
from functools import cache
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.security.http import HTTPAuthorizationCredentials, HTTPBearer

from dbgpt.component import ComponentType, SystemApp
from dbgpt.model.cluster import BaseModelController, WorkerManager, WorkerManagerFactory
from dbgpt_serve.core import Result
from dbgpt_serve.evaluate.api.schemas import EvaluateServeRequest
from dbgpt_serve.evaluate.config import SERVE_SERVICE_COMPONENT_NAME, ServeConfig
from dbgpt_serve.evaluate.service.service import Service

from ...prompt.service.service import Service as PromptService

router = APIRouter()

# Add your API endpoints here

global_system_app: Optional[SystemApp] = None
logger = logging.getLogger(__name__)


def get_service() -> Service:
    """Get the service instance"""
    return global_system_app.get_component(SERVE_SERVICE_COMPONENT_NAME, Service)


def get_prompt_service() -> PromptService:
    return global_system_app.get_component("dbgpt_serve_prompt_service", PromptService)


def get_worker_manager() -> WorkerManager:
    worker_manager = global_system_app.get_component(
        ComponentType.WORKER_MANAGER_FACTORY, WorkerManagerFactory
    ).create()
    return worker_manager


def get_model_controller() -> BaseModelController:
    controller = global_system_app.get_component(
        ComponentType.MODEL_CONTROLLER, BaseModelController
    )
    return controller


get_bearer_token = HTTPBearer(auto_error=False)


@cache
def _parse_api_keys(api_keys: str) -> List[str]:
    """Parse the string api keys to a list

    Args:
        api_keys (str): The string api keys

    Returns:
        List[str]: The list of api keys
    """
    if not api_keys:
        return []
    return [key.strip() for key in api_keys.split(",")]


async def check_api_key(
    auth: Optional[HTTPAuthorizationCredentials] = Depends(get_bearer_token),
    service: Service = Depends(get_service),
) -> Optional[str]:
    """Check the api key

    If the api key is not set, allow all.

    Your can pass the token in you request header like this:

    .. code-block:: python

        import requests

        client_api_key = "your_api_key"
        headers = {"Authorization": "Bearer " + client_api_key}
        res = requests.get("http://test/hello", headers=headers)
        assert res.status_code == 200

    """
    if service.config.api_keys:
        api_keys = _parse_api_keys(service.config.api_keys)
        if auth is None or (token := auth.credentials) not in api_keys:
            raise HTTPException(
                status_code=401,
                detail={
                    "error": {
                        "message": "",
                        "type": "invalid_request_error",
                        "param": None,
                        "code": "invalid_api_key",
                    }
                },
            )
        return token
    else:
        # api_keys not set; allow all
        return None


@router.get("/health", dependencies=[Depends(check_api_key)])
async def health():
    """Health check endpoint"""
    return {"status": "ok"}


@router.get("/test_auth", dependencies=[Depends(check_api_key)])
async def test_auth():
    """Test auth endpoint"""
    return {"status": "ok"}


@router.get("/scenes")
async def get_scenes():
    scene_list = [{"recall": "召回评测"}, {"app": "应用评测"}]

    return Result.succ(scene_list)


@router.get("/metrics")
async def get_metrics(scene_key: str, scene_value: str = None):
    """Get available evaluation metrics for a scene
    
    Args:
        scene_key (str): The scene key (e.g., 'app', 'recall')
        scene_value (str): The scene value (optional)
    
    Returns:
        Result: List of available metrics
    """
    metrics = []
    
    if scene_key == "recall":
        metrics = [
            {
                "name": "RetrieverSimilarityMetric",
                "description": "Embedding Similarity Metric for retrieval evaluation"
            },
            {
                "name": "RetrieverHitRateMetric", 
                "description": "Hit rate calculates the fraction of queries where the correct answer is found within the top-k retrieved documents"
            },
            {
                "name": "RetrieverMRRMetric",
                "description": "Mean Reciprocal Rank evaluates the system's accuracy by looking at the rank of the highest-placed relevant document"
            }
        ]
    elif scene_key == "app":
        metrics = [
            {
                "name": "AnswerRelevancyMetric",
                "description": "Evaluates the relevance of the answer to the given question"
            }
        ]
    
    return Result.succ(metrics)


@router.post("/evaluation")
async def evaluation(
    request: EvaluateServeRequest,
    service: Service = Depends(get_service),
) -> Result:
    """Evaluate results by the scene

    Args:
        request (EvaluateServeRequest): The request
        service (Service): The service
    Returns:
        ServerResponse: The response
    """
    return Result.succ(
        await service.run_evaluation(
            request.scene_key,
            request.scene_value,
            request.datasets,
            request.context,
            request.evaluate_metrics,
        )
    )


@router.delete("/evaluation")
async def delete_evaluation(
    evaluation_code: str,
    service: Service = Depends(get_service),
) -> Result:
    """Delete an evaluation by evaluation_code

    Args:
        evaluation_code (str): The evaluation code to delete
        service (Service): The service
    Returns:
        Result: The response indicating success or failure
    """
    try:
        # Check if evaluation exists
        evaluation = service.dao.get_one({"evaluate_code": evaluation_code})
        if not evaluation:
            return Result.failed(msg=f"Evaluation with code {evaluation_code} not found")
        
        # Delete the evaluation
        service.dao.delete({"evaluate_code": evaluation_code})
        
        return Result.succ({"message": f"Evaluation {evaluation_code} deleted successfully"})
    except Exception as e:
        return Result.failed(msg=f"Failed to delete evaluation: {str(e)}")


@router.post("/evaluations")
async def create_evaluation(
    request: EvaluateServeRequest,
    service: Service = Depends(get_service),
) -> Result:
    """Create and run a new evaluation, saving it to the database

    Args:
        request (EvaluateServeRequest): The request
        service (Service): The service
    Returns:
        ServerResponse: The response with the created evaluation
    """
    # Create and save the evaluation record
    evaluation_entity = service.create(request)
    
    # Run the evaluation and update the results
    try:
        results = await service.run_evaluation(
            request.scene_key,
            request.scene_value,
            request.datasets,
            request.context,
            request.evaluate_metrics,
        )
        
        # Calculate average score
        total_score = 0
        total_items = 0
        for result_list in results:
            for result in result_list:
                if result.score is not None:
                    total_score += float(result.score)
                    total_items += 1
        
        average_score = str(total_score / total_items) if total_items > 0 else None
        
        # Update the evaluation with results
        query_request = {"evaluate_code": evaluation_entity.evaluate_code}
        update_request = {
            "result": str(results),
            "state": "COMPLETE",
            "average_score": average_score
        }
        evaluation_entity = service.update(query_request, update_request)
        
    except Exception as e:
        # Update evaluation with error status
        query_request = {"evaluate_code": evaluation_entity.evaluate_code}
        update_request = {
            "state": "FAILED",
            "log_info": str(e)
        }
        evaluation_entity = service.update(query_request, update_request)
        raise
    
    return Result.succ(evaluation_entity)


@router.get("/evaluations")
async def get_evaluations(
    page: int = 1,
    page_size: int = 10,
    service: Service = Depends(get_service),
) -> Result:
    """Get all evaluations with pagination

    Args:
        page (int): Page number
        page_size (int): Page size
        service (Service): The service
    Returns:
        ServerResponse: The response with evaluations
    """
    evaluations = service.get_list_by_page({}, page, page_size)
    return Result.succ(evaluations)


@router.get("/datasets")
async def get_datasets(
    page: int = 1,
    page_size: int = 10,
    service: Service = Depends(get_service),
) -> Result:
    """Get all datasets with pagination

    Args:
        page (int): Page number
        page_size (int): Page size
        service (Service): The service
    Returns:
        Result: The response with datasets
    """
    # For now, return empty datasets list since we don't have a dataset management system
    # In a real implementation, this would fetch datasets from a dataset service
    datasets = {
        "items": [],
        "total_count": 0,
        "total_pages": 0,
        "page": page,
        "page_size": page_size
    }
    return Result.succ(datasets)


@router.get("/evaluation/detail/show")
async def get_evaluation_detail(
    evaluate_code: str,
    service: Service = Depends(get_service),
) -> Result:
    """Get evaluation detail by evaluate_code

    Args:
        evaluate_code (str): The evaluation code
        service (Service): The service
    Returns:
        Result: The response with evaluation detail
    """
    try:
        # Get the evaluation by evaluate_code
        evaluation = service.dao.get_one({"evaluate_code": evaluate_code})
        if not evaluation:
            return Result.failed(msg=f"Evaluation with code {evaluate_code} not found")
        
        return Result.succ(evaluation)
    except Exception as e:
        return Result.failed(msg=str(e))


@router.get("/evaluation/result/download")
async def download_evaluation_get(
    evaluate_code: str,
    service: Service = Depends(get_service),
):
    """Download evaluation results as a file (GET)

    Args:
        evaluate_code (str): The evaluation code
        service (Service): The service
    Returns:
        FileResponse: The evaluation results as a downloadable file
    """
    try:
        from fastapi.responses import Response
        import json
        from io import BytesIO
        
        # Get the evaluation by evaluate_code
        evaluation = service.dao.get_one({"evaluate_code": evaluate_code})
        if not evaluation:
            return Response(
                content=json.dumps({"error": f"Evaluation with code {evaluate_code} not found"}),
                status_code=404,
                media_type="application/json"
            )
        
        # Convert evaluation to Excel format
        evaluation_dict = evaluation.dict() if hasattr(evaluation, 'dict') else evaluation.__dict__
        
        try:
            # Try to use openpyxl for Excel generation
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Evaluation Report"
            
            # Set header style
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write headers
            ws['A1'] = 'Field'
            ws['B1'] = 'Value'
            ws['A1'].font = header_font
            ws['B1'].font = header_font
            ws['A1'].fill = header_fill
            ws['B1'].fill = header_fill
            
            row = 2
            
            # Write basic info
            basic_info = [
                ('Evaluation Code', evaluation_dict.get('evaluate_code', '')),
                ('Scene Key', evaluation_dict.get('scene_key', '')),
                ('Scene Value', evaluation_dict.get('scene_value', '')),
                ('Dataset Name', evaluation_dict.get('datasets_name', '')),
                ('State', evaluation_dict.get('state', '')),
                ('Average Score', evaluation_dict.get('average_score', '')),
                ('Created Time', str(evaluation_dict.get('gmt_create', ''))),
                ('Modified Time', str(evaluation_dict.get('gmt_modified', '')))
            ]
            
            for field, value in basic_info:
                ws[f'A{row}'] = field
                ws[f'B{row}'] = value
                row += 1
            
            # Write datasets
            datasets = evaluation_dict.get('datasets', [])
            if datasets:
                row += 1  # Empty row
                ws[f'A{row}'] = 'Datasets'
                ws[f'A{row}'].font = header_font
                row += 1
                
                for i, dataset in enumerate(datasets):
                    ws[f'A{row}'] = f'Dataset {i+1} Query'
                    ws[f'B{row}'] = dataset.get('query', '')
                    row += 1
                    ws[f'A{row}'] = f'Dataset {i+1} Doc Name'
                    ws[f'B{row}'] = dataset.get('doc_name', '')
                    row += 1
            
            # Write metrics
            metrics = evaluation_dict.get('evaluate_metrics', [])
            if metrics:
                row += 1  # Empty row
                ws[f'A{row}'] = 'Evaluation Metrics'
                ws[f'A{row}'].font = header_font
                ws[f'B{row}'] = ', '.join(metrics)
                row += 1
            
            # Write context
            context = evaluation_dict.get('context', {})
            if context:
                row += 1  # Empty row
                ws[f'A{row}'] = 'Context'
                ws[f'A{row}'].font = header_font
                row += 1
                
                for key, value in context.items():
                    ws[f'A{row}'] = f'Context {key}'
                    ws[f'B{row}'] = str(value)
                    row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_content = excel_buffer.getvalue()
            excel_buffer.close()
            
            # Return as Excel file
            return Response(
                content=excel_content,
                headers={
                    "Content-Disposition": f"attachment; filename=evaluation_{evaluate_code}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                }
            )
            
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            import csv
            from io import StringIO
            
            output = StringIO()
            writer = csv.writer(output)
            
            # Write CSV content (same as before)
            writer.writerow(['Field', 'Value'])
            writer.writerow(['Evaluation Code', evaluation_dict.get('evaluate_code', '')])
            writer.writerow(['Scene Key', evaluation_dict.get('scene_key', '')])
            writer.writerow(['Scene Value', evaluation_dict.get('scene_value', '')])
            writer.writerow(['Dataset Name', evaluation_dict.get('datasets_name', '')])
            writer.writerow(['State', evaluation_dict.get('state', '')])
            writer.writerow(['Average Score', evaluation_dict.get('average_score', '')])
            
            csv_content = output.getvalue()
            output.close()
            
            return Response(
                content=csv_content.encode('utf-8'),
                headers={
                    "Content-Disposition": f"attachment; filename=evaluation_{evaluate_code}.csv",
                    "Content-Type": "text/csv; charset=utf-8"
                }
            )
    except Exception as e:
        return Response(
            content=json.dumps({"error": str(e)}),
            status_code=500,
            media_type="application/json"
        )


def init_endpoints(system_app: SystemApp, config: ServeConfig) -> None:
    """Initialize the endpoints"""
    global global_system_app
    system_app.register(Service, config=config)
    global_system_app = system_app
